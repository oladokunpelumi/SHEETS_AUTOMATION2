"""
Twitter/X Profile Scanner v2
Enriched scanning: verification, followers, following, country/region, campaign-ready flag.

Usage:
  python twitter_scanner_v2.py input.xlsx --sheet "Sheet1" --output results.xlsx
  python twitter_scanner_v2.py input.xlsx --column B --headless --min-followers 1000
"""

import argparse
import json
import os
import re
import socket
import sys
import time
import random

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout


# ── Configuration ────────────────────────────────────────────────────────────
MIN_DELAY = 3
MAX_DELAY = 6
BATCH_PAUSE_EVERY = 50
BATCH_PAUSE_SECONDS = 30
PAGE_TIMEOUT = 20000
CHECKPOINT_EVERY = 25
MAX_RETRIES = 2
DEFAULT_MIN_FOLLOWERS = 1000  # Campaign-ready threshold


# ── Helper: Wait for internet connection ─────────────────────────────────────

def wait_for_internet():
    """Blocks and waits until an active internet connection is detected."""
    def is_connected():
        try:
            # Connect to Cloudflare DNS to check internet
            socket.create_connection(("1.1.1.1", 53), timeout=3)
            return True
        except OSError:
            pass
        return False

    if not is_connected():
        print("\n[!] Internet disconnected! Pausing scan... waiting for connection to resume.", flush=True)
        while not is_connected():
            time.sleep(5)
        print("[✓] Internet connection restored! Resuming scan...\n", flush=True)
        time.sleep(3)

def parse_follower_count(text: str) -> int | None:
    if not text:
        return None
    text = text.strip().split()[0].replace(",", "")
    multiplier = 1
    if text.upper().endswith("K"):
        multiplier = 1_000
        text = text[:-1]
    elif text.upper().endswith("M"):
        multiplier = 1_000_000
        text = text[:-1]
    elif text.upper().endswith("B"):
        multiplier = 1_000_000_000
        text = text[:-1]
    try:
        return int(float(text) * multiplier)
    except ValueError:
        return None


def format_count(count: int | None) -> str:
    if count is None:
        return "N/A"
    if count >= 1_000_000:
        return f"{count / 1_000_000:.1f}M"
    if count >= 1_000:
        return f"{count / 1_000:.1f}K"
    return str(count)


def extract_count(page, username: str, link_keyword: str) -> int | None:
    """Extract a count (followers or following) from a profile link."""
    selectors = [
        f'a[href="/{username}/verified_followers"] span span',
        f'a[href="/{username}/{link_keyword}"] span span',
        f'a[href="/{username}/{link_keyword}"] span',
    ] if link_keyword == "followers" else [
        f'a[href="/{username}/{link_keyword}"] span span',
        f'a[href="/{username}/{link_keyword}"] span',
    ]

    for sel in selectors:
        try:
            elements = page.locator(sel).all()
            for el in elements:
                text = el.inner_text(timeout=2000)
                if text and any(c.isdigit() for c in text):
                    count = parse_follower_count(text)
                    if count is not None:
                        return count
        except Exception:
            continue
    return None


def extract_location(page) -> str:
    """Extract the location/country from a Twitter profile."""
    location_selectors = [
        '[data-testid="UserProfileHeader_Items"] span[data-testid="UserLocation"]',
        '[data-testid="UserLocation"]',
    ]

    for sel in location_selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=1500):
                text = el.inner_text(timeout=2000).strip()
                if text:
                    return text
        except Exception:
            continue

    # Fallback: search page source for location pattern
    try:
        content = page.content()
        match = re.search(r'data-testid="UserLocation"[^>]*>([^<]+)<', content)
        if match:
            loc = match.group(1).strip()
            if loc:
                return loc
    except Exception:
        pass

    return ""


# ── Core: Scrape a single Twitter/X profile ──────────────────────────────────

def scrape_profile(page, username: str) -> dict:
    url = f"https://x.com/{username}"
    result = {
        "username": username,
        "verified": "No",
        "followers": "N/A",
        "followers_raw": None,
        "following": "N/A",
        "following_raw": None,
        "country_region": "",
        "status": "OK",
    }

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
        page.wait_for_timeout(3000)

        page_text = page.content()

        if "This account doesn't exist" in page_text or "Account suspended" in page_text:
            result["status"] = "Not Found / Suspended"
            return result

        if "These tweets are protected" in page_text:
            result["status"] = "Private Account"

        # ── Verified badge ────────────────────────────────────────────
        verified_selectors = [
            '[data-testid="icon-verified"]',
            'svg[aria-label="Verified account"]',
            'svg[aria-label="Verified"]',
            '[aria-label="Provides a verified blue checkmark"]',
            'a[href="/i/verified-choose"] svg',
        ]
        for sel in verified_selectors:
            try:
                if page.locator(sel).first.is_visible(timeout=1000):
                    result["verified"] = "Yes"
                    break
            except Exception:
                continue

        if result["verified"] == "No":
            if ('aria-label="Verified"' in page_text
                or 'data-testid="icon-verified"' in page_text
                or "Verified account" in page_text):
                result["verified"] = "Yes"

        # ── Follower count ────────────────────────────────────────────
        count = extract_count(page, username, "followers")
        if count is not None:
            result["followers_raw"] = count
            result["followers"] = format_count(count)
        else:
            match = re.search(r'([\d,.]+[KMB]?)\s*Followers', page_text)
            if match:
                count = parse_follower_count(match.group(1))
                if count is not None:
                    result["followers_raw"] = count
                    result["followers"] = format_count(count)

        # ── Following count ───────────────────────────────────────────
        count = extract_count(page, username, "following")
        if count is not None:
            result["following_raw"] = count
            result["following"] = format_count(count)
        else:
            match = re.search(r'([\d,.]+[KMB]?)\s*Following', page_text)
            if match:
                count = parse_follower_count(match.group(1))
                if count is not None:
                    result["following_raw"] = count
                    result["following"] = format_count(count)

        # ── Location / Country ────────────────────────────────────────
        result["country_region"] = extract_location(page)

    except PlaywrightTimeout:
        result["status"] = "Timeout"
    except Exception as e:
        result["status"] = f"Error: {str(e)[:60]}"

    return result


# ── Read usernames ───────────────────────────────────────────────────────────

def read_usernames(filepath: str, sheet_name: str = None, column: str = "A") -> list[str]:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb.active

    usernames = []
    col_idx = openpyxl.utils.column_index_from_string(column)
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
        val = row[0]
        if val and str(val).strip():
            usernames.append(str(val).strip().lstrip("@"))

    wb.close()
    print(f"Loaded {len(usernames)} usernames from '{filepath}' (sheet: {sheet_name or 'active'})")
    return usernames


# ── Write results to xlsx ────────────────────────────────────────────────────

def write_results(results: list[dict], output_path: str, min_followers: int):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Twitter Scan Results"

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1DA1F2")
    header_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    yes_fill = PatternFill("solid", fgColor="D4EDDA")
    no_fill = PatternFill("solid", fgColor="F8D7DA")
    error_fill = PatternFill("solid", fgColor="FFF3CD")
    ready_fill = PatternFill("solid", fgColor="C3E6CB")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    headers = [
        "#", "Username", "Blue Check Verified", "Followers", "Followers (Raw)",
        "Following", "Following (Raw)", "Follower/Following Ratio",
        "Country / Region", "Campaign Ready", "Status",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, r in enumerate(results, start=1):
        row = i + 1
        is_err = r["status"] not in ("OK", "Private Account")

        # Compute ratio
        ratio = ""
        if r["followers_raw"] and r["following_raw"] and r["following_raw"] > 0:
            ratio = f"{r['followers_raw'] / r['following_raw']:.1f}"

        # Campaign ready?
        campaign_ready = "No"
        if (r["followers_raw"] or 0) >= min_followers and r["status"] == "OK":
            campaign_ready = "Yes"

        cells_data = [
            (1, i, data_font, center),
            (2, f'@{r["username"]}', Font(name="Arial", size=10, color="1DA1F2"), left),
            (3, r["verified"], data_font, center),
            (4, r["followers"], data_font, right),
            (5, r["followers_raw"] if r["followers_raw"] else "N/A", data_font, right),
            (6, r["following"], data_font, right),
            (7, r["following_raw"] if r["following_raw"] else "N/A", data_font, right),
            (8, ratio if ratio else "N/A", data_font, center),
            (9, r["country_region"] if r["country_region"] else "—", data_font, left),
            (10, campaign_ready, data_font, center),
            (11, r["status"], data_font, center),
        ]

        for col, val, font, align in cells_data:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = font
            cell.alignment = align
            cell.border = thin_border

        # Conditional fills
        v_cell = ws.cell(row=row, column=3)
        if r["verified"] == "Yes":
            v_cell.fill = yes_fill
        elif not is_err:
            v_cell.fill = no_fill

        cr_cell = ws.cell(row=row, column=10)
        if campaign_ready == "Yes":
            cr_cell.fill = ready_fill
            cr_cell.font = Font(name="Arial", size=10, bold=True, color="155724")

        if is_err:
            ws.cell(row=row, column=11).fill = error_fill

    # Column widths
    widths = {"A": 6, "B": 25, "C": 22, "D": 14, "E": 16, "F": 14, "G": 16,
              "H": 22, "I": 28, "J": 16, "K": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{len(results) + 1}"

    # ── Summary sheet ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    total = len(results)
    verified = sum(1 for r in results if r["verified"] == "Yes")
    not_verified = sum(1 for r in results if r["verified"] == "No" and r["status"] == "OK")
    errors = sum(1 for r in results if r["status"] not in ("OK", "Private Account"))
    with_location = sum(1 for r in results if r["country_region"])
    campaign_count = sum(1 for r in results
                         if (r["followers_raw"] or 0) >= min_followers and r["status"] == "OK")

    summary = [
        ("Metric", "Value"),
        ("Total Usernames Scanned", total),
        ("Blue Check Verified (Yes)", verified),
        ("Not Verified (No)", not_verified),
        ("Errors / Not Found", errors),
        ("Verification Rate", f"{verified / max(total, 1) * 100:.1f}%"),
        ("Profiles with Location", with_location),
        (f"Campaign Ready ({min_followers}+ followers)", campaign_count),
    ]

    for row_idx, (label, value) in enumerate(summary, 1):
        a = ws2.cell(row=row_idx, column=1, value=label)
        b = ws2.cell(row=row_idx, column=2, value=value)
        a.border = thin_border
        b.border = thin_border
        if row_idx == 1:
            a.font = header_font; a.fill = header_fill
            b.font = header_font; b.fill = header_fill
        else:
            a.font = Font(name="Arial", size=10, bold=True)
            b.font = data_font; b.alignment = center

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 18

    # ── Campaign Ready sheet (filtered view) ──────────────────────────
    ws3 = wb.create_sheet("Campaign Ready")
    cr_headers = ["#", "Username", "Verified", "Followers", "Following", "Ratio", "Country / Region"]
    for col_idx, h in enumerate(cr_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="28A745")
        cell.alignment = header_align
        cell.border = thin_border

    cr_results = [r for r in results
                  if (r["followers_raw"] or 0) >= min_followers and r["status"] == "OK"]
    cr_results.sort(key=lambda x: x["followers_raw"] or 0, reverse=True)

    for i, r in enumerate(cr_results, start=1):
        row = i + 1
        ratio = ""
        if r["followers_raw"] and r["following_raw"] and r["following_raw"] > 0:
            ratio = f"{r['followers_raw'] / r['following_raw']:.1f}"

        data = [
            (1, i), (2, f'@{r["username"]}'), (3, r["verified"]),
            (4, r["followers"]), (5, r["following"]),
            (6, ratio if ratio else "N/A"),
            (7, r["country_region"] if r["country_region"] else "—"),
        ]
        for col, val in data:
            cell = ws3.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = center if col != 7 else left

    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 25
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 14
    ws3.column_dimensions["F"].width = 10
    ws3.column_dimensions["G"].width = 28
    ws3.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\nResults saved to: {output_path}")
    print(f"  Total: {total} | Verified: {verified} | Campaign Ready: {campaign_count} | Errors: {errors}")


# ── Checkpoint ───────────────────────────────────────────────────────────────

def save_checkpoint(results, path):
    with open(path, "w") as f:
        json.dump(results, f, indent=2)

def load_checkpoint(path):
    if os.path.exists(path):
        with open(path, "r") as f:
            return json.load(f)
    return []


# ── Main scan loop ───────────────────────────────────────────────────────────

def run_scan(input_file, sheet_name=None, column="A", output_file="results_v2.xlsx",
             headless=False, resume=True, min_followers=DEFAULT_MIN_FOLLOWERS):
    usernames = read_usernames(input_file, sheet_name, column)
    if not usernames:
        print("No usernames found.")
        return

    checkpoint_path = output_file.replace(".xlsx", "_checkpoint.json")
    results = load_checkpoint(checkpoint_path) if resume else []
    already_done = {r["username"].lower() for r in results}
    remaining = [u for u in usernames if u.lower() not in already_done]

    if not remaining:
        print("All done. Writing output...")
        write_results(results, output_file, min_followers)
        return

    print(f"\nRemaining: {len(remaining)} | Mode: {'Headless' if headless else 'Visible'}")
    print(f"Campaign-ready threshold: {min_followers}+ followers\n")

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 900},
            locale="en-US",
        )
        page = context.new_page()
        page.route("**/*", lambda route: route.abort() if route.request.resource_type in ("image", "media", "font") else route.continue_())

        scan_count = 0
        for i, username in enumerate(remaining):
            scan_count += 1
            progress = len(results) + 1
            total = len(usernames)
            pct = progress / total * 100

            print(f"  [{progress}/{total}] ({pct:.0f}%) @{username}...", end=" ", flush=True)

            result = None
            for attempt in range(1, MAX_RETRIES + 1):
                wait_for_internet()
                result = scrape_profile(page, username)
                if result["status"] in ("OK", "Private Account", "Not Found / Suspended"):
                    break
                
                # Check for explicit network loss before retry consumption
                if "net::ERR_INTERNET_DISCONNECTED" in result["status"] or result["status"] == "Timeout" or "Error" in result["status"]:
                    wait_for_internet()
                    
                if attempt < MAX_RETRIES:
                    print(f"(retry {attempt})...", end=" ", flush=True)
                    time.sleep(2)

            v = "✓" if result["verified"] == "Yes" else "✗"
            loc = result["country_region"][:20] if result["country_region"] else "—"
            print(f"{v} | {result['followers']} flw | {result['following']} flg | {loc} | {result['status']}")

            results.append(result)

            if scan_count % CHECKPOINT_EVERY == 0:
                save_checkpoint(results, checkpoint_path)
                print(f"    ↳ Checkpoint saved ({len(results)} scanned)")

            if scan_count % BATCH_PAUSE_EVERY == 0 and i < len(remaining) - 1:
                print(f"\n  ⏸  Batch pause ({BATCH_PAUSE_SECONDS}s)...\n")
                time.sleep(BATCH_PAUSE_SECONDS)
            else:
                time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

        browser.close()

    save_checkpoint(results, checkpoint_path)
    write_results(results, output_file, min_followers)
    if os.path.exists(checkpoint_path):
        os.remove(checkpoint_path)


def main():
    parser = argparse.ArgumentParser(description="Twitter/X Profile Scanner v2 (enriched)")
    parser.add_argument("input", help="Path to .xlsx file with usernames")
    parser.add_argument("--sheet", default=None, help="Sheet name")
    parser.add_argument("--column", default="A", help="Column with usernames (default: A)")
    parser.add_argument("--output", default="results_v2.xlsx", help="Output .xlsx file")
    parser.add_argument("--headless", action="store_true", help="Run browser headless")
    parser.add_argument("--no-resume", action="store_true", help="Ignore checkpoint")
    parser.add_argument("--min-followers", type=int, default=DEFAULT_MIN_FOLLOWERS,
                        help=f"Min followers for 'Campaign Ready' flag (default: {DEFAULT_MIN_FOLLOWERS})")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"Error: '{args.input}' not found.")
        sys.exit(1)

    run_scan(args.input, args.sheet, args.column, args.output, args.headless,
             not args.no_resume, args.min_followers)


if __name__ == "__main__":
    main()
