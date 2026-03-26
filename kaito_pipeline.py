"""
Kaito Yaps Leaderboard Scraper + Twitter/X Enrichment
Scrapes top creators from Kaito's yaps leaderboard, then enriches with Twitter data.

Two-step pipeline:
  Step 1: Scrape Kaito leaderboard → extract usernames + Yaps scores
  Step 2: Run each username through Twitter scanner → verified, followers, following, country

Usage:
  python kaito_pipeline.py --top 1000 --output kaito_creators.xlsx --headless

  # Step 1 only (just get Kaito usernames):
  python kaito_pipeline.py --top 1000 --kaito-only --output kaito_usernames.xlsx

  # Step 2 only (enrich already-scraped usernames from a previous run):
  python kaito_pipeline.py --enrich-from kaito_checkpoint.json --output kaito_creators.xlsx --headless
"""

import argparse
import json
import os
import re
import sys
import time
import random

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout


# ── Configuration ────────────────────────────────────────────────────────────
MIN_DELAY = 3
MAX_DELAY = 6
BATCH_PAUSE_EVERY = 50
BATCH_PAUSE_SECONDS = 45
PAGE_TIMEOUT = 20000
CHECKPOINT_EVERY = 25
MAX_RETRIES = 2
DEFAULT_MIN_FOLLOWERS = 1000

KAITO_LEADERBOARD_URL = "https://yaps.kaito.ai/leaderboard"
KAITO_API_URL = "https://api.kaito.ai/api/v1/yaps"


# ── Kaito API: get Yaps score for a single user ─────────────────────────────

def fetch_yaps_score(page, username: str) -> float | None:
    """Use Kaito's public API to get a user's Yaps score."""
    try:
        response = page.request.get(f"{KAITO_API_URL}?username={username}")
        if response.ok:
            data = response.json()
            return data.get("yaps_l30d", data.get("yaps_all", 0))
    except Exception:
        pass
    return None


# ── Kaito Leaderboard Scraper ────────────────────────────────────────────────

def scrape_kaito_leaderboard(target_count: int = 1000, headless: bool = False) -> list[dict]:
    """
    Scrape the Kaito Yaps leaderboard to get top creators.
    Returns list of dicts: {rank, username, yaps_score}
    """
    print(f"\n{'='*60}")
    print(f"  STEP 1: Scraping Kaito Leaderboard (top {target_count})")
    print(f"{'='*60}\n")

    creators = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 1000},
            locale="en-US",
        )
        page = context.new_page()

        print(f"  Loading {KAITO_LEADERBOARD_URL}...")
        page.goto(KAITO_LEADERBOARD_URL, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(5000)

        # Try to switch to a longer time range (30D or All) if available
        try:
            time_selectors = page.locator("text=30D").all()
            if time_selectors:
                time_selectors[0].click()
                page.wait_for_timeout(3000)
                print("  Switched to 30D view")
        except Exception:
            pass

        # Scroll and collect usernames
        seen_usernames = set()
        scroll_attempts = 0
        max_scroll_attempts = 200  # Safety limit
        stale_count = 0

        print(f"  Scrolling to load {target_count} creators...\n")

        while len(creators) < target_count and scroll_attempts < max_scroll_attempts:
            scroll_attempts += 1

            # Extract all visible username-like elements
            # Kaito displays X handles as @username in the leaderboard rows
            try:
                page_content = page.content()

                # Find all @username patterns in the page
                usernames_found = re.findall(r'@([A-Za-z0-9_]{1,15})', page_content)

                # Filter out common non-user handles
                skip = {"KaitoAI", "kaitoai", "x", "twitter", "dydx"}
                new_count = 0

                for uname in usernames_found:
                    if uname.lower() not in seen_usernames and uname not in skip and len(uname) > 1:
                        seen_usernames.add(uname.lower())
                        creators.append({
                            "rank": len(creators) + 1,
                            "username": uname,
                            "yaps_score": None,  # Will enrich via API if needed
                        })
                        new_count += 1

                if new_count == 0:
                    stale_count += 1
                else:
                    stale_count = 0

                if stale_count >= 15:
                    print(f"  No new creators after {stale_count} scrolls. Stopping at {len(creators)}.")
                    break

                if scroll_attempts % 10 == 0:
                    print(f"    Scroll #{scroll_attempts}: {len(creators)} creators found so far...")

            except Exception as e:
                print(f"    Error extracting: {e}")

            # Scroll down
            page.evaluate("window.scrollBy(0, 800)")
            page.wait_for_timeout(1500)

            # Occasional longer pause
            if scroll_attempts % 30 == 0:
                page.wait_for_timeout(3000)

        browser.close()

    # Trim to target
    creators = creators[:target_count]
    print(f"\n  Scraped {len(creators)} creators from Kaito leaderboard.")

    # ── Optional: Enrich Yaps scores via API ──────────────────────────
    # The public API has a 100 calls / 5 min rate limit.
    # For 1000 users this would take ~50 min. We skip this by default
    # and note it as optional.
    print("  (Yaps scores can be enriched via API — see --enrich-yaps flag)\n")

    return creators


def enrich_yaps_scores(creators: list[dict], headless: bool = False) -> list[dict]:
    """Enrich creator list with Yaps scores from Kaito API (rate limited)."""
    print(f"\n  Enriching {len(creators)} creators with Yaps scores via API...")
    print(f"  Rate limit: 100 calls / 5 min — this will take ~{len(creators) // 100 * 5} min\n")

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        for i, creator in enumerate(creators):
            if creator.get("yaps_score") is not None:
                continue

            score = fetch_yaps_score(page, creator["username"])
            creator["yaps_score"] = score

            if (i + 1) % 10 == 0:
                print(f"    [{i+1}/{len(creators)}] @{creator['username']} → {score}")

            # Rate limit: 100 per 5 min = ~1 per 3 seconds
            if (i + 1) % 95 == 0:
                print(f"    ⏸  API rate limit pause (60s)...")
                time.sleep(60)
            else:
                time.sleep(3.1)

        browser.close()

    return creators


# ── Twitter/X profile scraping (same as scanner v2) ─────────────────────────

def parse_follower_count(text):
    if not text:
        return None
    text = text.strip().split()[0].replace(",", "")
    mult = 1
    if text.upper().endswith("K"): mult = 1000; text = text[:-1]
    elif text.upper().endswith("M"): mult = 1_000_000; text = text[:-1]
    elif text.upper().endswith("B"): mult = 1_000_000_000; text = text[:-1]
    try:
        return int(float(text) * mult)
    except ValueError:
        return None

def format_count(c):
    if c is None: return "N/A"
    if c >= 1_000_000: return f"{c/1_000_000:.1f}M"
    if c >= 1_000: return f"{c/1_000:.1f}K"
    return str(c)

def extract_count(page, username, link_keyword):
    selectors = [
        f'a[href="/{username}/verified_followers"] span span',
        f'a[href="/{username}/{link_keyword}"] span span',
        f'a[href="/{username}/{link_keyword}"] span',
    ] if link_keyword == "followers" else [
        f'a[href="/{username}/{link_keyword}"] span span',
        f'a[href="/{username}/{link_keyword}"] span',
    ]
    for sel in selectors:
        try:
            for el in page.locator(sel).all():
                text = el.inner_text(timeout=2000)
                if text and any(c.isdigit() for c in text):
                    count = parse_follower_count(text)
                    if count is not None:
                        return count
        except Exception:
            continue
    return None

def extract_location(page):
    for sel in ['[data-testid="UserLocation"]']:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=1500):
                text = el.inner_text(timeout=2000).strip()
                if text:
                    return text
        except Exception:
            continue
    try:
        content = page.content()
        match = re.search(r'data-testid="UserLocation"[^>]*>([^<]+)<', content)
        if match:
            return match.group(1).strip()
    except Exception:
        pass
    return ""


def scrape_twitter_profile(page, username):
    url = f"https://x.com/{username}"
    result = {
        "username": username, "verified": "No",
        "followers": "N/A", "followers_raw": None,
        "following": "N/A", "following_raw": None,
        "country_region": "", "status": "OK",
    }

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
        page.wait_for_timeout(3000)
        page_text = page.content()

        if "This account doesn't exist" in page_text or "Account suspended" in page_text:
            result["status"] = "Not Found / Suspended"
            return result
        if "These tweets are protected" in page_text:
            result["status"] = "Private Account"

        # Verified
        for sel in ['[data-testid="icon-verified"]', 'svg[aria-label="Verified account"]',
                     'svg[aria-label="Verified"]', 'a[href="/i/verified-choose"] svg']:
            try:
                if page.locator(sel).first.is_visible(timeout=1000):
                    result["verified"] = "Yes"; break
            except Exception:
                continue
        if result["verified"] == "No" and ('aria-label="Verified"' in page_text or 'data-testid="icon-verified"' in page_text):
            result["verified"] = "Yes"

        # Followers
        c = extract_count(page, username, "followers")
        if c is not None:
            result["followers_raw"] = c; result["followers"] = format_count(c)
        else:
            m = re.search(r'([\d,.]+[KMB]?)\s*Followers', page_text)
            if m:
                c = parse_follower_count(m.group(1))
                if c: result["followers_raw"] = c; result["followers"] = format_count(c)

        # Following
        c = extract_count(page, username, "following")
        if c is not None:
            result["following_raw"] = c; result["following"] = format_count(c)
        else:
            m = re.search(r'([\d,.]+[KMB]?)\s*Following', page_text)
            if m:
                c = parse_follower_count(m.group(1))
                if c: result["following_raw"] = c; result["following"] = format_count(c)

        # Location
        result["country_region"] = extract_location(page)

    except PlaywrightTimeout:
        result["status"] = "Timeout"
    except Exception as e:
        result["status"] = f"Error: {str(e)[:60]}"

    return result


# ── Twitter enrichment loop ──────────────────────────────────────────────────

def enrich_with_twitter(creators: list[dict], headless: bool = False, resume: bool = True,
                         checkpoint_path: str = "kaito_twitter_checkpoint.json") -> list[dict]:
    print(f"\n{'='*60}")
    print(f"  STEP 2: Enriching {len(creators)} creators with Twitter data")
    print(f"{'='*60}\n")

    # Load checkpoint
    results = []
    if resume and os.path.exists(checkpoint_path):
        with open(checkpoint_path, "r") as f:
            results = json.load(f)
        print(f"  Resuming: {len(results)} already enriched.")

    already_done = {r["username"].lower() for r in results}
    remaining = [c for c in creators if c["username"].lower() not in already_done]

    if not remaining:
        print("  All creators already enriched.")
        return results

    print(f"  Remaining: {len(remaining)}\n")

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 900}, locale="en-US",
        )
        page = context.new_page()
        page.route("**/*", lambda route: route.abort() if route.request.resource_type in ("image", "media", "font") else route.continue_())

        scan_count = 0
        for i, creator in enumerate(remaining):
            scan_count += 1
            progress = len(results) + 1
            total = len(creators)
            pct = progress / total * 100
            username = creator["username"]

            print(f"  [{progress}/{total}] ({pct:.0f}%) @{username}...", end=" ", flush=True)

            tw = None
            for attempt in range(1, MAX_RETRIES + 1):
                tw = scrape_twitter_profile(page, username)
                if tw["status"] in ("OK", "Private Account", "Not Found / Suspended"):
                    break
                if attempt < MAX_RETRIES:
                    print(f"(retry)...", end=" ", flush=True)
                    time.sleep(2)

            # Merge Kaito + Twitter data
            merged = {
                "rank": creator.get("rank", ""),
                "username": username,
                "yaps_score": creator.get("yaps_score"),
                "verified": tw["verified"],
                "followers": tw["followers"],
                "followers_raw": tw["followers_raw"],
                "following": tw["following"],
                "following_raw": tw["following_raw"],
                "country_region": tw["country_region"],
                "status": tw["status"],
            }

            v = "✓" if tw["verified"] == "Yes" else "✗"
            loc = tw["country_region"][:20] if tw["country_region"] else "—"
            print(f"{v} | {tw['followers']} | {loc} | {tw['status']}")

            results.append(merged)

            if scan_count % CHECKPOINT_EVERY == 0:
                with open(checkpoint_path, "w") as f:
                    json.dump(results, f, indent=2)
                print(f"    ↳ Checkpoint saved ({len(results)})")

            if scan_count % BATCH_PAUSE_EVERY == 0 and i < len(remaining) - 1:
                print(f"\n  ⏸  Batch pause ({BATCH_PAUSE_SECONDS}s)...\n")
                time.sleep(BATCH_PAUSE_SECONDS)
            else:
                time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

        browser.close()

    # Final checkpoint
    with open(checkpoint_path, "w") as f:
        json.dump(results, f, indent=2)

    return results


# ── Write Kaito results to xlsx ──────────────────────────────────────────────

def write_kaito_results(results: list[dict], output_path: str, min_followers: int):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kaito Creators"

    hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    kaito_fill = PatternFill("solid", fgColor="7C3AED")  # Purple for Kaito brand
    ha = Alignment(horizontal="center", vertical="center")
    df = Font(name="Arial", size=10)
    c_al = Alignment(horizontal="center", vertical="center")
    l_al = Alignment(horizontal="left", vertical="center")
    r_al = Alignment(horizontal="right", vertical="center")
    yes_fill = PatternFill("solid", fgColor="D4EDDA")
    no_fill = PatternFill("solid", fgColor="F8D7DA")
    err_fill = PatternFill("solid", fgColor="FFF3CD")
    ready_fill = PatternFill("solid", fgColor="C3E6CB")
    brd = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
                 top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))

    headers = ["Rank", "Username", "Yaps Score", "Verified", "Followers", "Followers (Raw)",
               "Following", "Ratio", "Country / Region", "Campaign Ready", "Status"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hf; cell.fill = kaito_fill; cell.alignment = ha; cell.border = brd

    for i, r in enumerate(results, start=1):
        row = i + 1
        is_err = r["status"] not in ("OK", "Private Account")
        ratio = ""
        if r.get("followers_raw") and r.get("following_raw") and r["following_raw"] > 0:
            ratio = f"{r['followers_raw'] / r['following_raw']:.1f}"
        campaign = "Yes" if (r.get("followers_raw") or 0) >= min_followers and r["status"] == "OK" else "No"

        vals = [
            (1, r.get("rank", i), df, c_al),
            (2, f'@{r["username"]}', Font(name="Arial", size=10, color="1DA1F2"), l_al),
            (3, round(r["yaps_score"], 1) if r.get("yaps_score") else "N/A", df, r_al),
            (4, r["verified"], df, c_al),
            (5, r["followers"], df, r_al),
            (6, r.get("followers_raw") or "N/A", df, r_al),
            (7, r["following"], df, r_al),
            (8, ratio or "N/A", df, c_al),
            (9, r["country_region"] or "—", df, l_al),
            (10, campaign, df, c_al),
            (11, r["status"], df, c_al),
        ]
        for col, val, font, align in vals:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = font; cell.alignment = align; cell.border = brd

        if r["verified"] == "Yes": ws.cell(row=row, column=4).fill = yes_fill
        elif not is_err: ws.cell(row=row, column=4).fill = no_fill
        if campaign == "Yes":
            ws.cell(row=row, column=10).fill = ready_fill
            ws.cell(row=row, column=10).font = Font(name="Arial", size=10, bold=True, color="155724")
        if is_err:
            ws.cell(row=row, column=11).fill = err_fill

    widths = {"A": 8, "B": 25, "C": 14, "D": 12, "E": 14, "F": 16, "G": 14,
              "H": 10, "I": 28, "J": 16, "K": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{len(results) + 1}"

    # ── Summary ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    total = len(results)
    verified = sum(1 for r in results if r["verified"] == "Yes")
    errors = sum(1 for r in results if r["status"] not in ("OK", "Private Account"))
    with_loc = sum(1 for r in results if r.get("country_region"))
    camp = sum(1 for r in results if (r.get("followers_raw") or 0) >= min_followers and r["status"] == "OK")

    summary = [
        ("Metric", "Value"),
        ("Total Creators Scraped", total),
        ("Blue Check Verified", verified),
        ("Errors / Not Found", errors),
        ("Profiles with Location", with_loc),
        (f"Campaign Ready ({min_followers}+ followers)", camp),
    ]
    for ri, (label, value) in enumerate(summary, 1):
        a = ws2.cell(row=ri, column=1, value=label)
        b = ws2.cell(row=ri, column=2, value=value)
        a.border = brd; b.border = brd
        if ri == 1:
            a.font = hf; a.fill = kaito_fill; b.font = hf; b.fill = kaito_fill
        else:
            a.font = Font(name="Arial", size=10, bold=True); b.font = df; b.alignment = c_al
    ws2.column_dimensions["A"].width = 38; ws2.column_dimensions["B"].width = 18

    # ── Campaign Ready sheet ──────────────────────────────────────────
    ws3 = wb.create_sheet("Campaign Ready")
    cr_headers = ["#", "Username", "Yaps", "Verified", "Followers", "Following", "Ratio", "Country"]
    for ci, h in enumerate(cr_headers, 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.font = hf; cell.fill = PatternFill("solid", fgColor="28A745")
        cell.alignment = ha; cell.border = brd

    cr = [r for r in results if (r.get("followers_raw") or 0) >= min_followers and r["status"] == "OK"]
    cr.sort(key=lambda x: x.get("followers_raw") or 0, reverse=True)

    for i, r in enumerate(cr, start=1):
        row = i + 1
        ratio = ""
        if r.get("followers_raw") and r.get("following_raw") and r["following_raw"] > 0:
            ratio = f"{r['followers_raw'] / r['following_raw']:.1f}"
        data = [(1, i), (2, f'@{r["username"]}'),
                (3, round(r["yaps_score"], 1) if r.get("yaps_score") else "N/A"),
                (4, r["verified"]), (5, r["followers"]), (6, r["following"]),
                (7, ratio or "N/A"), (8, r.get("country_region") or "—")]
        for col, val in data:
            cell = ws3.cell(row=row, column=col, value=val)
            cell.font = df; cell.border = brd; cell.alignment = c_al if col != 8 else l_al

    for col, w in {"A": 6, "B": 25, "C": 10, "D": 12, "E": 14, "F": 14, "G": 10, "H": 28}.items():
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\nResults saved to: {output_path}")
    print(f"  Total: {total} | Verified: {verified} | Campaign Ready: {camp} | Errors: {errors}")


# ── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Kaito Leaderboard → Twitter/X Enrichment Pipeline")
    parser.add_argument("--top", type=int, default=1000, help="How many top creators to scrape from Kaito (default: 1000)")
    parser.add_argument("--output", default="kaito_creators.xlsx", help="Output .xlsx file")
    parser.add_argument("--headless", action="store_true", help="Run browser headless")
    parser.add_argument("--kaito-only", action="store_true", help="Only scrape Kaito (skip Twitter enrichment)")
    parser.add_argument("--enrich-from", default=None, help="Path to a Kaito checkpoint JSON to enrich (skip Kaito scrape)")
    parser.add_argument("--enrich-yaps", action="store_true", help="Also fetch Yaps scores via API (slow, rate limited)")
    parser.add_argument("--min-followers", type=int, default=DEFAULT_MIN_FOLLOWERS, help="Campaign-ready threshold")
    parser.add_argument("--no-resume", action="store_true", help="Ignore checkpoints, start fresh")
    args = parser.parse_args()

    checkpoint_path = args.output.replace(".xlsx", "_twitter_checkpoint.json")

    # ── Step 1: Get Kaito creators ────────────────────────────────────
    if args.enrich_from:
        print(f"Loading creators from: {args.enrich_from}")
        with open(args.enrich_from, "r") as f:
            creators = json.load(f)
        print(f"  Loaded {len(creators)} creators.")
    else:
        creators = scrape_kaito_leaderboard(target_count=args.top, headless=args.headless)

        # Save Kaito scrape
        kaito_checkpoint = args.output.replace(".xlsx", "_kaito_checkpoint.json")
        with open(kaito_checkpoint, "w") as f:
            json.dump(creators, f, indent=2)
        print(f"  Kaito checkpoint saved: {kaito_checkpoint}")

        if args.enrich_yaps:
            creators = enrich_yaps_scores(creators, headless=True)
            with open(kaito_checkpoint, "w") as f:
                json.dump(creators, f, indent=2)

    if args.kaito_only:
        # Write Kaito-only results (no Twitter data)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kaito Creators"
        ws.append(["Rank", "Username", "Yaps Score"])
        for c in creators:
            ws.append([c.get("rank", ""), c["username"], c.get("yaps_score", "N/A")])
        wb.save(args.output)
        print(f"\nKaito-only results saved to: {args.output}")
        return

    # ── Step 2: Enrich with Twitter data ──────────────────────────────
    results = enrich_with_twitter(
        creators, headless=args.headless,
        resume=not args.no_resume, checkpoint_path=checkpoint_path,
    )

    # ── Step 3: Write final output ────────────────────────────────────
    write_kaito_results(results, args.output, args.min_followers)

    # Clean up
    if os.path.exists(checkpoint_path):
        os.remove(checkpoint_path)
        print("Checkpoint cleaned up.")


if __name__ == "__main__":
    main()
